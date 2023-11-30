import boto3
import openpyxl
from openpyxl.styles import Font, Alignment
import configparser
from datetime import datetime
from dateutil.relativedelta import relativedelta

# Read the config.ini file
config = configparser.ConfigParser()
config.read('config.ini')

# Read the AWS region from config.ini
region = config.get('AWS', 'region')

# Read the tags from config.ini
tags = {}
for tag_key, tag_value in config.items('TAGS'):
    if tag_value.strip():
        tags[tag_key.strip()] = tag_value.strip()

# Create a new workbook and add a worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Format column headers
header_font = Font(bold=True)
worksheet["A1"].font = header_font
worksheet["B1"].font = header_font
worksheet["C1"].font = header_font
worksheet["D1"].font = header_font
worksheet["E1"].font = header_font
worksheet["F1"].font = header_font

# Set column widths
worksheet.column_dimensions["A"].width = 20
worksheet.column_dimensions["B"].width = 15
worksheet.column_dimensions["C"].width = 12
worksheet.column_dimensions["D"].width = 12
worksheet.column_dimensions["E"].width = 15
worksheet.column_dimensions["F"].width = 20

# Set header row values
worksheet["A1"] = "Instance ID"
worksheet["B1"] = "Instance Type"
worksheet["C1"] = "Max Value"
worksheet["D1"] = "Min Value"
worksheet["E1"] = "Avg Value"
worksheet["F1"] = "Tag Values"

# Center-align header row cells
for cell in worksheet[1]:
    cell.alignment = Alignment(horizontal="center")

# Create a CloudWatch client
cloudwatch = boto3.client("cloudwatch", region_name=region)

# Create an EC2 client
ec2 = boto3.client("ec2", region_name=region)

# Get a list of all running EC2 instances in the region
response = ec2.describe_instances(Filters=[{"Name": "instance-state-name", "Values": ["running"]}])

instances = []
for reservation in response["Reservations"]:
    for instance in reservation["Instances"]:
        instances.append(instance)

# Calculate the start time and end time for the duration of 4 months
end_time = datetime.utcnow()
start_time = end_time - relativedelta(months=4)

# Loop through each instance and retrieve the metric statistics
row = 2
for instance in instances:
    instance_id = instance["InstanceId"]
    instance_type = instance["InstanceType"]
    print(f"Processing instance {instance_id}...")

    # Get the metric statistics (max, min, avg) for the specified duration
    metric_data = cloudwatch.get_metric_statistics(
        Namespace="AWS/EC2",
        MetricName="CPUUtilization",
        StartTime=start_time,
        EndTime=end_time,
        Period=10800,
        Statistics=["Maximum", "Minimum", "Average"],
        Dimensions=[{"Name": "InstanceId", "Value": instance_id}]
    )

    # Extract the metric values
    data_points = metric_data["Datapoints"]
    if data_points:
        max_value = max(data_points, key=lambda x: x["Maximum"])["Maximum"]
        min_value = min(data_points, key=lambda x: x["Minimum"])["Minimum"]
        avg_value = sum([point["Average"] for point in data_points]) / len(data_points)
    else:
        max_value = min_value = avg_value = None

    # Write the instance details to the worksheet
    worksheet.cell(row=row, column=1, value=instance_id)
    worksheet.cell(row=row, column=2, value=instance_type)
    worksheet.cell(row=row, column=3, value=max_value)
    worksheet.cell(row=row, column=4, value=min_value)
    worksheet.cell(row=row, column=5, value=avg_value)

    # Fetch and write the tag values to the worksheet
    tags_values = []
    for tag_key, tag_value in tags.items():
        if tag_value.strip():
            tags_values.append(f"{tag_key}: {tag_value.strip()}")
        else:
            ec2_resource = boto3.resource("ec2", region_name=region)
            ec2_instance = ec2_resource.Instance(instance_id)
            aws_tags = ec2_instance.tags
            for aws_tag in aws_tags:
                if aws_tag["Key"] == tag_key:
                    tag_value = aws_tag["Value"]
                    break
            tags_values.append(f"{tag_key}: {tag_value}")

    worksheet.cell(row=row, column=6, value=", ".join(tags_values))

    row += 1

# Save the workbook
workbook.save("instance_metrics.xlsx")
print("Excel file generated successfully!")

